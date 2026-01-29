import base64
import io
import json
import logging
import re
import uuid
from concurrent.futures import ThreadPoolExecutor

import streamlit as st
from bs4 import BeautifulSoup
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from streamlit_autorefresh import st_autorefresh

logger = logging.getLogger("translate_app")
if not logger.handlers:
    handler = logging.StreamHandler()
    handler.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] [%(threadName)s] %(message)s")
    )
    handler.setLevel(logging.INFO)
    logger.addHandler(handler)
logger.setLevel(logging.INFO)
logger.propagate = False


def build_client(settings):
    api_key = settings["api_key"]
    base_url = settings.get("base_url")
    if base_url:
        return OpenAI(api_key=api_key, base_url=base_url)
    return OpenAI(api_key=api_key)


@st.cache_resource
def get_executor():
    return ThreadPoolExecutor(max_workers=10)


def extract_json(text):
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(json)?", "", cleaned)
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except Exception:
        pass
    match = re.search(r"\{.*\}", cleaned, flags=re.S)
    if match:
        return json.loads(match.group(0))
    match = re.search(r"\[.*\]", cleaned, flags=re.S)
    if match:
        return json.loads(match.group(0))
    raise ValueError("无法解析模型输出的 JSON")


def should_skip_cell(value):
    """判断单元格是否应该跳过不翻译"""
    if value is None:
        return True
    if not isinstance(value, str):
        return True
    stripped = value.strip()
    if not stripped:
        return True
    # 跳过公式
    if stripped.startswith("="):
        return True
    return False


def sheet_to_grid(ws):
    """将 worksheet 转为二维数组，只保留有内容的区域（去掉尾部空行空列）"""
    # 先找出实际有内容的范围
    max_row = 0
    max_col = 0
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        for col_idx, cell in enumerate(row, start=1):
            if not should_skip_cell(cell.value):
                max_row = max(max_row, row_idx)
                max_col = max(max_col, col_idx)

    if max_row == 0 or max_col == 0:
        return []

    # 只读取有内容的区域
    grid = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        row_data = []
        for cell in row:
            if should_skip_cell(cell.value):
                row_data.append(None)
            else:
                row_data.append(cell.value)
        grid.append(row_data)
    return grid


def grid_to_sheet(ws, translated_grid):
    """将翻译后的二维数组写回 worksheet"""
    for row_idx, row_data in enumerate(translated_grid, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                cell = ws.cell(row=row_idx, column=col_idx)
                # 只更新非公式单元格
                if not should_skip_cell(cell.value):
                    cell.value = value


def translate_grid(grid, source_lang, target_lang, model, api_settings, task_id=None):
    """翻译二维表格，保留上下文"""
    # 统计非空单元格数量
    cell_count = sum(1 for row in grid for cell in row if cell is not None)
    if cell_count == 0:
        return grid

    logger.info(
        "translate_grid_start task_id=%s rows=%s cols=%s cells=%s",
        task_id,
        len(grid),
        max(len(row) for row in grid) if grid else 0,
        cell_count,
    )

    client = build_client(api_settings)
    system_prompt = (
        "你是专业的外贸/工业品翻译引擎，擅长机械、电气、五金、化工等领域术语。\n"
        "输入是表格的二维数组（JSON），请翻译后返回相同结构的二维数组。\n"
        "规则：\n"
        "1. 使用正式商务语体，术语准确，表达简洁\n"
        "2. 根据上下文（同行/同列内容）理解每个单元格的含义，相同文字可能有不同译法\n"
        "3. 保持原样不翻译：型号、规格、参数值、认证标准、品牌名、货号、单位符号\n"
        "4. 保持原样不翻译：纯数字、日期、URL、邮箱、文件路径\n"
        "5. 保留原文的换行符、空格、大小写和标点风格\n"
        "6. null 值保持为 null\n"
        "只输出 JSON 数组，禁止任何解释或 Markdown。"
    )
    user_prompt = (
        f"源语言: {source_lang}\n目标语言: {target_lang}\n"
        "请翻译以下表格：\n"
        f"{json.dumps(grid, ensure_ascii=False)}"
    )

    logger.info(
        "translate_grid_request task_id=%s model=%s base_url=%s prompt_len=%s",
        task_id,
        model,
        api_settings.get("base_url"),
        len(user_prompt),
    )

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        temperature=0,
    )
    content = response.choices[0].message.content or ""
    logger.info(
        "translate_grid_response task_id=%s content_len=%s",
        task_id,
        len(content),
    )

    try:
        translated = extract_json(content)
        if not isinstance(translated, list):
            raise ValueError("模型返回的不是数组")
    except Exception:
        logger.exception(
            "translate_grid_parse_error task_id=%s content_head=%s",
            task_id,
            content[:300].replace("\n", "\\n"),
        )
        raise

    logger.info(
        "translate_grid_done task_id=%s translated_rows=%s",
        task_id,
        len(translated),
    )
    return translated


def translate_workbook_in_place(wb, source_lang, target_lang, model, api_settings, task_id=None):
    """翻译整个工作簿，按 sheet 处理"""
    for sheet_idx, ws in enumerate(wb.worksheets):
        sheet_task_id = f"{task_id}-sheet{sheet_idx}" if task_id else None
        logger.info(
            "translate_sheet_start task_id=%s sheet=%s rows=%s cols=%s",
            sheet_task_id,
            ws.title,
            ws.max_row,
            ws.max_column,
        )

        # 读取为二维数组
        grid = sheet_to_grid(ws)

        # 检查是否有内容需要翻译
        has_content = any(cell is not None for row in grid for cell in row)
        if not has_content:
            logger.info("translate_sheet_skip task_id=%s sheet=%s (empty)", sheet_task_id, ws.title)
            continue

        # 翻译
        translated_grid = translate_grid(
            grid,
            source_lang,
            target_lang,
            model,
            api_settings,
            task_id=sheet_task_id,
        )

        # 写回
        grid_to_sheet(ws, translated_grid)
        logger.info("translate_sheet_done task_id=%s sheet=%s", sheet_task_id, ws.title)

    return wb


def image_to_html_table(image_bytes, model, api_settings, task_id=None, table_rows=None, table_cols=None):
    client = build_client(api_settings)
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    logger.info(
        "image_vision_start task_id=%s bytes=%s model=%s base_url=%s rows=%s cols=%s",
        task_id,
        len(image_bytes),
        model,
        api_settings.get("base_url"),
        table_rows,
        table_cols,
    )

    # 根据是否提供行列数，构建不同的 prompt
    if table_rows and table_cols:
        system_prompt = (
            f"识别图片中的表格，该表格为 {table_rows} 行 {table_cols} 列。\n"
            f"请严格按照 {table_rows} 行 {table_cols} 列的结构输出 HTML <table>。\n"
            "规则：\n"
            "1. 原样输出所有文字内容，不要翻译或修改任何文字\n"
            f"2. 每行必须有且仅有 {table_cols} 个 <td> 或 <th>\n"
            "3. 空白单元格必须输出为空 <td></td>\n"
            "4. 单元格内的换行用 <br>\n"
            "5. 不要添加 style/class 等属性\n"
            "6. 不要输出 Markdown、解释或多余文本"
        )
    else:
        system_prompt = (
            "识别图片中的表格结构与内容，仅输出一个完整的 HTML <table>。\n"
            "规则：\n"
            "1. 原样输出所有文字内容，不要翻译或修改任何文字\n"
            "2. 必须用 <tr>/<td>/<th> 表达结构，必要时使用 rowspan/colspan\n"
            "3. 空白单元格也要输出为空 <td></td>，保持列数一致\n"
            "4. 单元格内的换行用 <br>\n"
            "5. 不要添加 style/class 等属性\n"
            "6. 不要输出 Markdown、解释或多余文本"
        )

    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "请处理这张图片。"},
                    {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}},
                ],
            },
        ],
        temperature=0,
    )
    content = response.choices[0].message.content or ""
    logger.info(
        "image_vision_response task_id=%s content_len=%s",
        task_id,
        len(content),
    )
    return content


def html_to_excel_with_format(html_str):
    soup = BeautifulSoup(html_str, "lxml")
    table = soup.find("table")
    if table is None:
        raise ValueError("未找到 table 标签")
    wb = Workbook()
    ws = wb.active
    occupied = set()
    row_idx = 1
    for tr in table.find_all("tr"):
        col_idx = 1
        cells = tr.find_all(["td", "th"])
        for cell in cells:
            while (row_idx, col_idx) in occupied:
                col_idx += 1
            text = cell.get_text(separator="\n", strip=True)
            rowspan = int(cell.get("rowspan", 1))
            colspan = int(cell.get("colspan", 1))
            ws.cell(row=row_idx, column=col_idx, value=text)
            if rowspan > 1 or colspan > 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx + rowspan - 1,
                    end_column=col_idx + colspan - 1,
                )
            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    if r == row_idx and c == col_idx:
                        continue
                    occupied.add((r, c))
            col_idx += colspan
        row_idx += 1
    return wb


def run_excel_translate_job(file_bytes, source_lang, target_lang, model, api_settings, task_id=None):
    logger.info(
        "excel_job_start task_id=%s bytes=%s",
        task_id,
        len(file_bytes),
    )
    wb = load_workbook(io.BytesIO(file_bytes))
    wb = translate_workbook_in_place(
        wb,
        source_lang,
        target_lang,
        model,
        api_settings,
        task_id=task_id,
    )
    output = io.BytesIO()
    wb.save(output)
    logger.info(
        "excel_job_done task_id=%s output_bytes=%s",
        task_id,
        output.tell(),
    )
    return output.getvalue()


def run_image_vision_job(image_bytes, vision_model, vision_api_settings, task_id=None, table_rows=None, table_cols=None):
    return image_to_html_table(
        image_bytes,
        model=vision_model,
        api_settings=vision_api_settings,
        task_id=task_id,
        table_rows=table_rows,
        table_cols=table_cols,
    )


def run_workbook_translate_job(
    workbook_bytes,
    source_lang,
    target_lang,
    model,
    api_settings,
    task_id=None,
):
    logger.info(
        "workbook_job_start task_id=%s bytes=%s",
        task_id,
        len(workbook_bytes),
    )
    wb = load_workbook(io.BytesIO(workbook_bytes))
    wb = translate_workbook_in_place(
        wb,
        source_lang,
        target_lang,
        model,
        api_settings,
        task_id=task_id,
    )
    output = io.BytesIO()
    wb.save(output)
    logger.info(
        "workbook_job_done task_id=%s output_bytes=%s",
        task_id,
        output.tell(),
    )
    return output.getvalue()


def require_login():
    if st.session_state.get("authenticated"):
        return
    st.title("Login")
    username = st.text_input("用户名")
    password = st.text_input("密码", type="password")
    if st.button("登录"):
        passwords = st.secrets.get("passwords", {})
        if username in passwords and passwords[username] == password:
            st.session_state["authenticated"] = True
            st.session_state["username"] = username
            st.rerun()
        else:
            st.error("用户名或密码错误")
    st.stop()


def main():
    st.set_page_config(page_title="zyt's translator")
    # 隐藏右上角三个点菜单
    st.markdown(
        """
        <style>
        #MainMenu {visibility: hidden;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    require_login()
    with st.sidebar:
        st.write(f"当前用户: {st.session_state.get('username', '')}")
        source_lang = st.text_input("源语言", value="Auto")
        target_lang = st.text_input("目标语言", value="Chinese")
        if st.button("退出登录"):
            st.session_state.clear()
            st.rerun()
    st.title("zyt's translator")
    text_api_settings = st.secrets.get("text_api_settings")
    vision_api_settings = st.secrets.get("vision_api_settings")
    if not text_api_settings:
        st.error("未配置 text_api_settings")
        st.stop()
    if not vision_api_settings:
        st.error("未配置 vision_api_settings")
        st.stop()
    models = st.secrets.get("models", {})
    text_model = models.get("text_model")
    vision_model = models.get("vision_model")
    if not text_model:
        st.error("未配置 models.text_model")
        st.stop()
    if not vision_model:
        st.error("未配置 models.vision_model")
        st.stop()
    text_api_settings = dict(text_api_settings)
    vision_api_settings = dict(vision_api_settings)
    executor = get_executor()
    running_tasks = False
    tab_excel, tab_image = st.tabs(["Excel 翻译", "图片翻译"])
    with tab_excel:
        # 检查任务状态
        excel_future = st.session_state.get("excel_future")
        excel_is_running = False
        if excel_future:
            if excel_future.done():
                try:
                    st.session_state["excel_result"] = excel_future.result()
                    st.session_state["excel_status"] = "完成"
                    st.session_state["excel_error"] = None
                except Exception as exc:
                    logger.exception(
                        "excel_future_error task_id=%s",
                        st.session_state.get("excel_task_id"),
                    )
                    st.session_state["excel_result"] = None
                    st.session_state["excel_status"] = None
                    st.session_state["excel_error"] = str(exc)
                finally:
                    st.session_state["excel_future"] = None
            else:
                st.session_state["excel_status"] = "正在翻译"
                running_tasks = True
                excel_is_running = True

        excel_status = st.session_state.get("excel_status")
        excel_error = st.session_state.get("excel_error")
        excel_result = st.session_state.get("excel_result")

        # 状态机：根据当前状态决定显示什么
        # 1. 翻译中 -> 只显示状态条
        # 2. 完成 -> 显示完成状态 + 下载按钮
        # 3. 其他（初始/错误）-> 显示表单

        if excel_is_running:
            # 翻译中：只显示状态条
            st.info("正在翻译，请稍候...")
        elif excel_status == "完成" and excel_result:
            # 完成：显示成功状态和下载按钮
            st.success("翻译完成！")

            def reset_excel_state():
                st.session_state["excel_status"] = None
                st.session_state["excel_error"] = None
                st.session_state["excel_result"] = None
                st.session_state["excel_future"] = None

            st.download_button(
                "📥 下载翻译结果",
                data=excel_result,
                file_name="translated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="excel_download",
                on_click=reset_excel_state,
            )
        else:
            # 初始状态或错误状态：显示表单
            if excel_error:
                st.error(excel_error)

            with st.form("excel_form"):
                excel_file = st.file_uploader("上传 .xlsx 文件", type=["xlsx"])
                model = st.text_input("模型", value=text_model)
                submitted = st.form_submit_button("开始翻译")

            if excel_file and submitted:
                excel_bytes = excel_file.getvalue()
                # 防御：检查文件是否为空
                if len(excel_bytes) == 0:
                    st.session_state["excel_error"] = "上传的文件为空，请重新选择文件"
                    st.rerun()
                # 防御：检查文件大小（限制 50MB）
                elif len(excel_bytes) > 50 * 1024 * 1024:
                    st.session_state["excel_error"] = "文件过大（超过 50MB），请拆分后上传"
                    st.rerun()
                else:
                    task_id = f"excel-{uuid.uuid4().hex[:8]}"
                    st.session_state["excel_task_id"] = task_id
                    logger.info(
                        "excel_submit task_id=%s user=%s bytes=%s model=%s source=%s target=%s",
                        task_id,
                        st.session_state.get("username"),
                        len(excel_bytes),
                        model,
                        source_lang,
                        target_lang,
                    )
                    st.session_state["excel_status"] = "正在翻译"
                    st.session_state["excel_error"] = None
                    st.session_state["excel_result"] = None
                    st.session_state["excel_future"] = executor.submit(
                        run_excel_translate_job,
                        excel_bytes,
                        source_lang,
                        target_lang,
                        model,
                        text_api_settings,
                        task_id,
                    )
                    running_tasks = True
                    st.rerun()
    with tab_image:
        # 检查任务状态
        vision_future = st.session_state.get("image_vision_future")
        image_is_running = False
        image_current_step = None  # 用于显示当前步骤

        if vision_future:
            if vision_future.done():
                try:
                    html_str = vision_future.result()
                    task_id = st.session_state.get("image_task_id")
                    logger.info(
                        "image_vision_done task_id=%s html_len=%s has_table=%s head=%s",
                        task_id,
                        len(html_str),
                        "<table" in html_str.lower(),
                        html_str[:200].replace("\n", "\\n"),
                    )
                    wb = html_to_excel_with_format(html_str)
                    ws = wb.active
                    logger.info(
                        "image_table_parsed task_id=%s rows=%s cols=%s",
                        task_id,
                        ws.max_row,
                        ws.max_column,
                    )
                    buf = io.BytesIO()
                    wb.save(buf)
                    workbook_bytes = buf.getvalue()
                    logger.info(
                        "image_translate_submit task_id=%s bytes=%s model=%s source=%s target=%s",
                        task_id,
                        len(workbook_bytes),
                        st.session_state.get("image_translate_model"),
                        source_lang,
                        target_lang,
                    )
                    translate_task_id = f"{task_id}-translate" if task_id else None
                    st.session_state["image_translate_future"] = executor.submit(
                        run_workbook_translate_job,
                        workbook_bytes,
                        source_lang,
                        target_lang,
                        st.session_state.get("image_translate_model"),
                        text_api_settings,
                        translate_task_id,
                    )
                    st.session_state["image_vision_future"] = None
                    st.session_state["image_status"] = "正在翻译"
                    st.session_state["image_error"] = None
                    image_is_running = True
                    image_current_step = "正在翻译"  # 统一翻译提示
                except Exception as exc:
                    logger.exception(
                        "image_vision_error task_id=%s",
                        st.session_state.get("image_task_id"),
                    )
                    st.session_state["image_translate_future"] = None
                    st.session_state["image_vision_future"] = None
                    st.session_state["image_status"] = None
                    st.session_state["image_error"] = str(exc)
            else:
                st.session_state["image_status"] = "正在识别"
                running_tasks = True
                image_is_running = True
                image_current_step = "正在识别图片中的表格"

        translate_future = st.session_state.get("image_translate_future")
        if translate_future:
            if translate_future.done():
                try:
                    st.session_state["image_result"] = translate_future.result()
                    st.session_state["image_status"] = "完成"
                    st.session_state["image_error"] = None
                except Exception as exc:
                    logger.exception(
                        "image_translate_error task_id=%s",
                        st.session_state.get("image_task_id"),
                    )
                    st.session_state["image_result"] = None
                    st.session_state["image_status"] = None
                    st.session_state["image_error"] = str(exc)
                finally:
                    st.session_state["image_translate_future"] = None
            else:
                st.session_state["image_status"] = "正在翻译"
                running_tasks = True
                image_is_running = True
                image_current_step = "正在翻译"  # 统一翻译提示

        image_status = st.session_state.get("image_status")
        image_error = st.session_state.get("image_error")
        image_result = st.session_state.get("image_result")

        # 状态机：根据当前状态决定显示什么
        # 1. 处理中 -> 只显示状态条
        # 2. 完成 -> 显示完成状态 + 下载按钮
        # 3. 其他（初始/错误）-> 显示表单

        if image_is_running:
            # 处理中：只显示状态条
            st.info(f"{image_current_step}，请稍候...")
        elif image_status == "完成" and image_result:
            # 完成：显示成功状态和下载按钮
            st.success("翻译完成！")

            def reset_image_state():
                st.session_state["image_status"] = None
                st.session_state["image_error"] = None
                st.session_state["image_result"] = None
                st.session_state["image_vision_future"] = None
                st.session_state["image_translate_future"] = None

            st.download_button(
                "📥 下载翻译结果",
                data=image_result,
                file_name="table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="image_download",
                on_click=reset_image_state,
            )
        else:
            # 初始状态或错误状态：显示表单
            if image_error:
                st.error(image_error)

            with st.form("image_form"):
                image_file = st.file_uploader("上传表格图片", type=["png", "jpg", "jpeg", "webp"])
                col1, col2 = st.columns(2)
                with col1:
                    table_rows = st.number_input(
                        "表格行数（可选，填写可提高识别准确度）",
                        min_value=0,
                        max_value=1000,
                        value=0,
                        help="填 0 表示自动识别",
                    )
                with col2:
                    table_cols = st.number_input(
                        "表格列数（可选，填写可提高识别准确度）",
                        min_value=0,
                        max_value=100,
                        value=0,
                        help="填 0 表示自动识别",
                    )
                vision_model_input = st.text_input("视觉模型", value=vision_model)
                translate_model = st.text_input("翻译模型", value=text_model)
                submitted = st.form_submit_button("开始翻译")

            if image_file and submitted:
                image_bytes = image_file.getvalue()
                # 防御：检查文件是否为空
                if len(image_bytes) == 0:
                    st.session_state["image_error"] = "上传的文件为空，请重新选择文件"
                    st.rerun()
                # 防御：检查文件大小（限制 20MB）
                elif len(image_bytes) > 20 * 1024 * 1024:
                    st.session_state["image_error"] = "图片过大（超过 20MB），请压缩后上传"
                    st.rerun()
                else:
                    task_id = f"image-{uuid.uuid4().hex[:8]}"
                    st.session_state["image_task_id"] = task_id
                    # 0 表示自动识别，转为 None
                    rows_hint = table_rows if table_rows > 0 else None
                    cols_hint = table_cols if table_cols > 0 else None
                    logger.info(
                        "image_submit task_id=%s user=%s bytes=%s vision_model=%s translate_model=%s rows=%s cols=%s",
                        task_id,
                        st.session_state.get("username"),
                        len(image_bytes),
                        vision_model_input,
                        translate_model,
                        rows_hint,
                        cols_hint,
                    )
                    st.session_state["image_translate_model"] = translate_model
                    st.session_state["image_status"] = "正在识别"
                    st.session_state["image_error"] = None
                    st.session_state["image_result"] = None
                    st.session_state["image_vision_future"] = executor.submit(
                        run_image_vision_job,
                        image_bytes,
                        vision_model_input,
                        vision_api_settings,
                        f"{task_id}-vision",
                        rows_hint,
                        cols_hint,
                    )
                    running_tasks = True
                    st.rerun()
    if running_tasks:
        st_autorefresh(interval=1500, limit=None, key="task_refresh")


if __name__ == "__main__":
    main()
